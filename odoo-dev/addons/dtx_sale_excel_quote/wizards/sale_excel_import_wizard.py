# -*- coding: utf-8 -*-
import logging
import base64
from io import BytesIO
from odoo import models, fields, api, _
from odoo.exceptions import UserError, ValidationError

_logger = logging.getLogger(__name__)

try:
    from openpyxl import load_workbook
except ImportError:
    _logger.warning('openpyxl library not found. Excel import will not work.')


class SaleExcelImportWizard(models.TransientModel):
    _name = 'sale.excel.import.wizard'
    _description = 'Wizard to import Sale Quotation from Excel'

    order_id = fields.Many2one(
        'sale.order',
        string='Sale Order',
        required=True,
        readonly=True,
    )

    file = fields.Binary(
        string='File Excel',
        required=True,
        help='Upload file Excel theo template DTX',
    )

    filename = fields.Char(
        string='Tên file',
    )

    import_mode = fields.Selection([
        ('replace', 'Thay thế toàn bộ (xóa lines cũ)'),
        ('append', 'Thêm vào (giữ lines cũ)'),
    ], string='Chế độ import', default='replace', required=True)

    def action_import(self):
        """Import Excel file and update sale order lines"""
        self.ensure_one()

        if not self.file:
            raise UserError('Vui lòng upload file Excel.')

        if self.order_id.state not in ['draft', 'sent']:
            raise UserError('Chỉ có thể import vào quotation ở trạng thái Draft hoặc Sent.')

        try:
            # Decode file
            file_data = base64.b64decode(self.file)
            wb = load_workbook(BytesIO(file_data), data_only=False)

            # Get first sheet
            ws = wb.active

            # Parse and create lines
            new_lines = self._parse_excel_lines(ws)

            # Apply import mode
            if self.import_mode == 'replace':
                # Delete existing lines
                self.order_id.order_line.unlink()

            # Create new lines
            for line_vals in new_lines:
                line_vals['order_id'] = self.order_id.id
                self.env['sale.order.line'].create(line_vals)

            # Show success message
            return {
                'type': 'ir.actions.client',
                'tag': 'display_notification',
                'params': {
                    'title': _('Thành công'),
                    'message': f'Đã import {len(new_lines)} dòng vào báo giá {self.order_id.name}',
                    'type': 'success',
                    'sticky': False,
                }
            }

        except Exception as e:
            _logger.error(f"Error importing Excel for order {self.order_id.name}: {str(e)}", exc_info=True)
            raise UserError(f'Lỗi khi import file Excel:\n{str(e)}')

    def _parse_excel_lines(self, ws):
        """Parse Excel worksheet and return list of line values"""
        lines = []
        sequence = 10

        # Find header row (contains "STT" and "MÔ TẢ SẢN PHẨM")
        header_row = None
        for row_idx in range(1, min(30, ws.max_row + 1)):
            cell_a = ws.cell(row=row_idx, column=1).value
            cell_b = ws.cell(row=row_idx, column=2).value

            if cell_a and cell_b:
                if 'STT' in str(cell_a).upper() and 'MÔ TẢ' in str(cell_b).upper():
                    header_row = row_idx
                    break

        if not header_row:
            raise ValidationError(
                'Không tìm thấy header table trong file Excel.\n'
                'Header phải có dòng chứa: STT | MÔ TẢ SẢN PHẨM | ...'
            )

        _logger.info(f"Found header row at: {header_row}")

        # Parse data rows (start from header + 1)
        data_start_row = header_row + 1

        for row_idx in range(data_start_row, ws.max_row + 1):
            # Read columns
            col_a = ws.cell(row=row_idx, column=1).value  # STT
            col_b = ws.cell(row=row_idx, column=2).value  # MÔ TẢ
            col_c = ws.cell(row=row_idx, column=3).value  # MÃ SP
            col_d = ws.cell(row=row_idx, column=4).value  # XUẤT XỨ
            col_e = ws.cell(row=row_idx, column=5).value  # ĐVT
            col_f = ws.cell(row=row_idx, column=6).value  # SL
            col_g = ws.cell(row=row_idx, column=7).value  # ĐƠN GIÁ
            col_h = ws.cell(row=row_idx, column=8).value  # VAT

            # Skip empty rows
            if not col_b and not col_c:
                continue

            # Check for totals row (stop parsing)
            if col_b and any(x in str(col_b).lower() for x in ['tổng cộng', 'vat', 'thành tiền']):
                _logger.info(f"Reached totals section at row {row_idx}, stopping parse.")
                break

            # Determine line type
            line_vals = {'sequence': sequence}

            # Section line: has col_b but no product code and no qty/price
            if col_b and not col_c and not col_f and not col_g:
                # Check if it's a section (bold/header style) or note
                # For simplicity, if col_a has roman numerals or letters like "I.", "A.", consider it section
                if col_a and self._is_section_marker(col_a):
                    line_vals['display_type'] = 'line_section'
                    line_vals['name'] = f"{col_a} {col_b}" if col_a else col_b
                else:
                    # Note line
                    line_vals['display_type'] = 'line_note'
                    line_vals['name'] = col_b

            # Product line: has product code or has qty/price
            elif col_c or col_f or col_g:
                # Find product by code
                product = None
                if col_c:
                    product = self.env['product.product'].search([
                        ('default_code', '=', str(col_c).strip())
                    ], limit=1)

                    if not product:
                        # Try fuzzy search by name
                        product = self.env['product.product'].search([
                            ('name', 'ilike', str(col_c).strip())
                        ], limit=1)

                if not product:
                    raise ValidationError(
                        f'Không tìm thấy sản phẩm với mã "{col_c}" tại dòng {row_idx}.\n'
                        f'Vui lòng kiểm tra mã sản phẩm trong hệ thống.'
                    )

                line_vals['product_id'] = product.id
                line_vals['name'] = col_b or product.display_name

                # Quantity
                try:
                    qty = float(col_f) if col_f else 1.0
                    if qty <= 0:
                        raise ValueError("Quantity must be positive")
                    line_vals['product_uom_qty'] = qty
                except (ValueError, TypeError) as e:
                    raise ValidationError(
                        f'Số lượng không hợp lệ tại dòng {row_idx}: "{col_f}"\n'
                        f'Lỗi: {str(e)}'
                    )

                # UOM
                if col_e:
                    uom = self.env['uom.uom'].search([
                        ('name', '=', str(col_e).strip())
                    ], limit=1)
                    if uom:
                        line_vals['product_uom'] = uom.id
                    else:
                        line_vals['product_uom'] = product.uom_id.id
                else:
                    line_vals['product_uom'] = product.uom_id.id

                # Price
                try:
                    price = float(col_g) if col_g else 0.0
                    line_vals['price_unit'] = price
                except (ValueError, TypeError) as e:
                    raise ValidationError(
                        f'Đơn giá không hợp lệ tại dòng {row_idx}: "{col_g}"\n'
                        f'Lỗi: {str(e)}'
                    )

                # VAT - map to tax_ids
                vat_rate = self._extract_vat_rate(ws, row_idx, col_g, col_h)
                if vat_rate is not None:
                    tax = self._find_tax_by_rate(vat_rate)
                    if tax:
                        line_vals['tax_id'] = [(6, 0, [tax.id])]
                    else:
                        _logger.warning(f"No tax found for VAT rate {vat_rate}% at row {row_idx}")

            else:
                # Unknown line type, skip
                _logger.warning(f"Skipping unknown line type at row {row_idx}")
                continue

            lines.append(line_vals)
            sequence += 10

        _logger.info(f"Parsed {len(lines)} lines from Excel")
        return lines

    def _is_section_marker(self, value):
        """Check if value is a section marker like I., A., B., etc."""
        if not value:
            return False

        s = str(value).strip().upper()

        # Roman numerals with dot
        if s in ['I.', 'II.', 'III.', 'IV.', 'V.', 'VI.', 'VII.', 'VIII.', 'IX.', 'X.']:
            return True

        # Single letters with dot
        if len(s) == 2 and s[0].isalpha() and s[1] == '.':
            return True

        # Numbers with dot (could be section)
        if s.replace('.', '').isdigit() and '.' in s:
            return True

        return False

    def _extract_vat_rate(self, ws, row_idx, unit_price_val, vat_val):
        """Extract VAT rate from Excel cell (formula or value)"""
        try:
            # Get cell H (VAT column)
            vat_cell = ws.cell(row=row_idx, column=8)

            # Check if it's a formula
            if vat_cell.value and isinstance(vat_cell.value, str) and vat_cell.value.startswith('='):
                formula = vat_cell.value
                # Parse formula like "=G16*8%" or "=G16*10%"
                if '*' in formula and '%' in formula:
                    parts = formula.split('*')
                    if len(parts) == 2:
                        rate_str = parts[1].replace('%', '').strip()
                        try:
                            return float(rate_str)
                        except ValueError:
                            pass

            # If not formula, calculate from values
            if vat_val and unit_price_val:
                try:
                    vat_amount = float(vat_val)
                    unit_price = float(unit_price_val)

                    if unit_price > 0:
                        vat_rate = (vat_amount / unit_price) * 100
                        # Round to nearest standard VAT rate
                        if abs(vat_rate - 0) < 0.5:
                            return 0.0
                        elif abs(vat_rate - 8) < 0.5:
                            return 8.0
                        elif abs(vat_rate - 10) < 0.5:
                            return 10.0
                        else:
                            return round(vat_rate, 2)
                except (ValueError, TypeError, ZeroDivisionError):
                    pass

            return None

        except Exception as e:
            _logger.warning(f"Error extracting VAT rate at row {row_idx}: {str(e)}")
            return None

    def _find_tax_by_rate(self, rate):
        """Find account.tax by rate (0%, 8%, 10%)"""
        # Search for sales tax with matching rate
        tax = self.env['account.tax'].search([
            ('type_tax_use', '=', 'sale'),
            ('amount_type', '=', 'percent'),
            ('amount', '=', rate),
            ('company_id', '=', self.order_id.company_id.id),
        ], limit=1)

        return tax
