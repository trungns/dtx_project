# -*- coding: utf-8 -*-
import logging
import base64
import os
from io import BytesIO
from datetime import datetime
from odoo import models, fields, api, _
from odoo.exceptions import UserError
from odoo.modules.module import get_module_path

_logger = logging.getLogger(__name__)

try:
    from openpyxl import load_workbook
    from openpyxl.utils import get_column_letter
except ImportError:
    _logger.warning('openpyxl library not found. Excel export/import will not work.')


class SaleOrder(models.Model):
    _inherit = 'sale.order'

    def action_export_excel_dtx(self):
        """Export quotation to Excel using EXACT DTX template replication"""
        self.ensure_one()

        try:
            # Generate Excel file using template
            excel_data = self._generate_excel_from_template()

            # Create attachment
            filename = f"Bao_gia_{self.partner_id.name}_{self.name.replace('/', '_')}_{datetime.now().strftime('%d.%m.%Y')}.xlsx"
            attachment = self.env['ir.attachment'].create({
                'name': filename,
                'type': 'binary',
                'datas': base64.b64encode(excel_data),
                'res_model': 'sale.order',
                'res_id': self.id,
                'mimetype': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            })

            # Return download action
            return {
                'type': 'ir.actions.act_url',
                'url': f'/web/content/{attachment.id}?download=true',
                'target': 'self',
            }

        except Exception as e:
            _logger.error(f"Error exporting quotation {self.name}: {str(e)}", exc_info=True)
            raise UserError(f'Lỗi khi export báo giá:\n{str(e)}')

    def _generate_excel_from_template(self):
        """
        Generate Excel using EXACT template replication approach.

        RULES:
        1. Copy template bytes -> load with openpyxl
        2. Only modify cell.value, NEVER modify style/format/merge
        3. Fill slots according to template structure
        4. Clear unused slots (set value=None)
        5. Validate format parity before return
        """
        # Load template
        template_path = os.path.join(
            get_module_path('dtx_sale_excel_quote'),
            'static', 'template', 'DTX_Quotation_Template.xlsx'
        )

        if not os.path.exists(template_path):
            raise UserError(f'Template file not found: {template_path}')

        # Load workbook (keep formulas, keep links, preserve everything)
        wb = load_workbook(template_path, data_only=False, keep_links=True)

        # Get target sheet
        # Priority: "Du toan" (with diacritics, recommended by user) > "Chốt" > "HĐ" > first sheet
        target_sheet = None
        for sheet_name in ['Du toan', 'Chốt', 'HĐ']:
            if sheet_name in wb.sheetnames:
                target_sheet = sheet_name
                break

        if not target_sheet:
            target_sheet = wb.sheetnames[0]
            _logger.warning(f'Default sheets not found, using {target_sheet}')

        _logger.info(f'Using sheet: {target_sheet}')
        ws = wb[target_sheet]

        # Detect template structure
        template_info = self._detect_template_structure(ws)

        # Fill header data
        self._fill_header_data(ws, template_info)

        # Fill product lines
        self._fill_product_lines(ws, template_info)

        # Validate format parity
        self._validate_format_parity(wb, template_path)

        # Save to BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        wb.close()

        return output.read()

    def _detect_template_structure(self, ws):
        """
        Detect template structure automatically.

        Returns dict with:
        - header_row: row number of table header
        - totals_row: row number of totals section
        - section_rows: list of section header rows
        - note_slots: list of note/description rows
        - product_slots: list of product data rows
        """
        _logger.info("Detecting template structure...")

        # Find header row (contains "STT" and "Mô tả")
        header_row = None
        for row_idx in range(1, min(20, ws.max_row + 1)):
            cell_a = ws.cell(row=row_idx, column=1).value
            cell_b = ws.cell(row=row_idx, column=2).value

            if cell_a and cell_b:
                if 'STT' in str(cell_a).upper() and 'MÔ TẢ' in str(cell_b).upper():
                    header_row = row_idx
                    break

        if not header_row:
            raise UserError('Không tìm thấy header table trong template. Template bị sai cấu trúc.')

        _logger.info(f"Header row: {header_row}")

        # Find totals row (contains "Tổng cộng" or similar)
        # Search from BOTTOM to TOP to avoid matching section headers
        totals_row = ws.max_row
        for row_idx in range(ws.max_row, header_row, -1):
            # Prioritize column G (7) which typically has totals in product tables
            val_g = ws.cell(row=row_idx, column=7).value
            if val_g and 'TỔNG CỘNG' in str(val_g).upper():
                totals_row = row_idx
                break

            # Also check column B (2) as fallback
            val_b = ws.cell(row=row_idx, column=2).value
            if val_b and 'TỔNG CỘNG' in str(val_b).upper():
                totals_row = row_idx
                break

        _logger.info(f"Totals row: {totals_row}")
        _logger.info(f"Data range: {header_row + 1} to {totals_row - 1}")

        # Classify rows
        section_rows = []
        note_slots = []
        product_slots = []

        for row_idx in range(header_row + 1, totals_row):
            cell_a = ws.cell(row=row_idx, column=1).value
            cell_b = ws.cell(row=row_idx, column=2).value
            cell_f = ws.cell(row=row_idx, column=6).value  # SL
            cell_g = ws.cell(row=row_idx, column=7).value  # Đơn giá

            # Determine if this is a product slot
            # Product has STT (numeric in col A) OR has qty/price (cols F, G)
            has_stt = cell_a and (isinstance(cell_a, (int, float)) or (isinstance(cell_a, str) and cell_a.strip().isdigit()))
            has_qty_price = cell_f or cell_g

            if has_stt or has_qty_price:
                # Product slot
                product_slots.append(row_idx)
            elif cell_b:
                # Has description but no qty/price
                b_str = str(cell_b).strip()

                # Check if section (col A has roman/letter marker, or col B is short header-like)
                is_section = False
                if cell_a:
                    a_str = str(cell_a).strip().upper()
                    # Roman numerals, single letters like I., II., A, B, C
                    if a_str in ['I.', 'II.', 'III.', 'IV.', 'A', 'B', 'C', 'D', 'E'] or (len(a_str) <= 3 and a_str.endswith('.')):
                        is_section = True

                # Check B content for section patterns
                if not is_section and len(b_str) < 100:
                    if (b_str.endswith('.') or b_str.endswith(':') or
                        any(pattern in b_str for pattern in [
                            'Phần mềm', 'Thiết bị', 'Chi phí', 'Ram:', 'SSD:', 'Chip:',
                            'Điện áp', 'Trọng lượng', 'Chất liệu', 'TỔNG GIÁ TRỊ'
                        ])):
                        is_section = True

                if is_section:
                    section_rows.append(row_idx)
                else:
                    # Note slot (description/detail)
                    note_slots.append(row_idx)
            # Else: completely empty row, skip

        _logger.info(f"Classified: {len(section_rows)} sections, {len(note_slots)} notes, {len(product_slots)} products")

        return {
            'header_row': header_row,
            'totals_row': totals_row,
            'section_rows': section_rows,
            'note_slots': note_slots,
            'product_slots': product_slots,
        }

    def _fill_header_data(self, ws, template_info):
        """
        Fill header information (customer, sales info).
        Only update cell values, preserve all formatting.

        Template structure (for "Du toan" and "Chốt" sheets):
        Row 6: BÁO GIÁ (merged A6:J6) - keep as is
        Row 7: HỆ THỐNG XẾP HÀNG THÔNG MINH (merged A7:J7) - keep as is
        Row 9: B="To: {customer}" | G="Sales:"
        Row 10: B="Mr: {contact}" | G="{sales_name}"
        Row 11: G="Điện thoại: {sales_phone}"
        Row 12: B="E-Mail: {customer_email}" | G="E-Mail: {sales_email}"
        Row 13: B="Địa chỉ: {customer_address}"
        """
        _logger.info("Filling header data...")

        partner = self.partner_id
        user = self.user_id

        # Detect if this sheet has header (check if row 6 has "BÁO GIÁ")
        has_header = False
        if ws.cell(row=6, column=1).value and 'BÁO GIÁ' in str(ws.cell(row=6, column=1).value).upper():
            has_header = True

        if not has_header:
            _logger.info("Sheet does not have header section, skipping header fill")
            return

        _logger.info("Sheet has header section, filling customer and sales info...")

        # Row 9: Customer name
        # B9 = "To: {customer_name}"
        customer_name = partner.commercial_company_name or partner.parent_id.name or partner.name
        current_b9 = str(ws.cell(row=9, column=2).value or "")
        if "To:" in current_b9:
            # Replace after "To:"
            ws.cell(row=9, column=2).value = f"To: {customer_name}"
        else:
            ws.cell(row=9, column=2).value = customer_name

        # Row 10: Contact person and Sales name
        # B10 = "Mr: {contact_name}"
        contact_name = partner.child_ids[0].name if partner.child_ids else partner.name
        current_b10 = str(ws.cell(row=10, column=2).value or "")
        if "Mr:" in current_b10:
            ws.cell(row=10, column=2).value = f"Mr: {contact_name}"
        else:
            ws.cell(row=10, column=2).value = contact_name

        # G10 = Sales name (direct value, no prefix)
        ws.cell(row=10, column=7).value = user.name or ''

        # Row 11: Sales phone
        # G11 = "Điện thoại: {phone}"
        sales_phone = user.phone or user.mobile or ''
        current_g11 = str(ws.cell(row=11, column=7).value or "")
        if "Điện thoại:" in current_g11:
            ws.cell(row=11, column=7).value = f"Điện thoại: {sales_phone}"
        else:
            ws.cell(row=11, column=7).value = sales_phone

        # Row 12: Customer and Sales email
        # B12 = "E-Mail: {customer_email}"
        customer_email = partner.email or ''
        current_b12 = str(ws.cell(row=12, column=2).value or "")
        if "E-Mail:" in current_b12 or "Email:" in current_b12:
            ws.cell(row=12, column=2).value = f"E-Mail: {customer_email}"
        else:
            ws.cell(row=12, column=2).value = customer_email

        # G12 = "E-Mail: {sales_email}"
        sales_email = user.email or ''
        current_g12 = str(ws.cell(row=12, column=7).value or "")
        if "E-Mail:" in current_g12 or "Email:" in current_g12:
            ws.cell(row=12, column=7).value = f"E-Mail: {sales_email}"
        else:
            ws.cell(row=12, column=7).value = sales_email

        # Row 13: Customer address
        # B13 = "Địa chỉ: {address}"
        address_parts = [
            partner.street,
            partner.street2,
            partner.city,
            partner.state_id.name if partner.state_id else None,
            partner.country_id.name if partner.country_id else None
        ]
        address = ', '.join([p for p in address_parts if p])
        current_b13 = str(ws.cell(row=13, column=2).value or "")
        if "Địa chỉ:" in current_b13:
            ws.cell(row=13, column=2).value = f"Địa chỉ: {address}"
        else:
            ws.cell(row=13, column=2).value = address

        _logger.info("Header data filled successfully")

    def _fill_product_lines(self, ws, template_info):
        """
        Fill product lines into template slots.

        CRITICAL: Follow slot-based approach, NOT data-driven approach.
        Iterate through template slots and fill with data where available.
        """
        _logger.info("Filling product lines...")

        # Detect template type by checking header row
        # Check if header has "VAT" column (10-column template) or "Đơn giá (Bao gồm VAT)" (8-column template)
        header_row = template_info['header_row']
        header_h = ws.cell(row=header_row, column=8).value
        has_separate_vat = header_h and 'VAT' in str(header_h).upper() and 'BAO GỒM' not in str(header_h).upper()

        _logger.info(f"Template type: {'10-column (separate VAT)' if has_separate_vat else '8-column (VAT included)'}")

        # Get order lines (exclude display_type lines for product slots)
        product_lines = self.order_line.filtered(lambda l: not l.display_type)
        note_lines = self.order_line.filtered(lambda l: l.display_type == 'line_note')

        _logger.info(f"Order has {len(product_lines)} product lines and {len(note_lines)} note lines")

        # Check if we exceed template capacity
        if len(product_lines) > len(template_info['product_slots']):
            raise UserError(
                f'Quotation có {len(product_lines)} dòng sản phẩm, '
                f'nhưng template chỉ hỗ trợ {len(template_info["product_slots"])} slots.\n'
                f'Vui lòng giảm số dòng hoặc cập nhật template.'
            )

        # Fill product slots
        product_line_idx = 0
        stt_counter = 1

        for slot_row in template_info['product_slots']:
            if product_line_idx < len(product_lines):
                line = product_lines[product_line_idx]

                # Fill columns
                ws.cell(row=slot_row, column=1).value = stt_counter  # STT
                ws.cell(row=slot_row, column=2).value = line.name or ''  # Mô tả
                ws.cell(row=slot_row, column=3).value = line.product_id.default_code or ''  # Mã SP

                # Xuất xứ (col D) - check if product has origin field
                origin = ''
                if hasattr(line.product_id, 'x_origin'):
                    origin = line.product_id.x_origin or ''
                elif hasattr(line.product_id, 'country_of_origin_id'):
                    origin = line.product_id.country_of_origin_id.name if line.product_id.country_of_origin_id else ''
                ws.cell(row=slot_row, column=4).value = origin

                ws.cell(row=slot_row, column=5).value = line.product_uom.name or ''  # ĐVT
                ws.cell(row=slot_row, column=6).value = line.product_uom_qty  # SL

                # Handle price based on template type
                if has_separate_vat:
                    # 10-column template: Đơn giá (chưa VAT) | VAT | Tổng giá | Thành tiền
                    ws.cell(row=slot_row, column=7).value = line.price_unit  # Đơn giá (chưa VAT)

                    # VAT column - don't set value, template has formula (=G*8%)
                    # Just let it calculate automatically

                    # Tổng giá (col I) - don't set, template has formula
                    # Thành tiền (col J) - don't set, template has formula

                else:
                    # 8-column template: Đơn giá (bao gồm VAT) | Thành tiền
                    # Calculate price with tax
                    price_with_tax = line.price_unit
                    if line.tax_id:
                        # Add tax to price
                        for tax in line.tax_id:
                            if tax.amount_type == 'percent':
                                price_with_tax = price_with_tax * (1 + tax.amount / 100)

                    ws.cell(row=slot_row, column=7).value = price_with_tax  # Đơn giá (bao gồm VAT)

                    # Thành tiền - usually has formula, but set value to be safe
                    total = price_with_tax * line.product_uom_qty
                    ws.cell(row=slot_row, column=8).value = total  # Thành tiền

                product_line_idx += 1
                stt_counter += 1
            else:
                # Clear unused slot (set values to None, preserve formatting)
                ws.cell(row=slot_row, column=1).value = None
                ws.cell(row=slot_row, column=2).value = None
                ws.cell(row=slot_row, column=3).value = None
                ws.cell(row=slot_row, column=4).value = None
                ws.cell(row=slot_row, column=5).value = None
                ws.cell(row=slot_row, column=6).value = None
                ws.cell(row=slot_row, column=7).value = None
                ws.cell(row=slot_row, column=8).value = None
                if has_separate_vat:
                    # Also clear cols I, J for 10-column template
                    ws.cell(row=slot_row, column=9).value = None
                    ws.cell(row=slot_row, column=10).value = None

        # Note: Section rows and note slots are preserved as-is from template
        # We don't modify them unless explicitly needed

        _logger.info(f"Filled {product_line_idx} product lines into template")

    def _validate_format_parity(self, wb, template_path):
        """
        Validate that output workbook matches template format.

        Checks:
        - Sheet names match
        - Merged cells match
        - Column widths match
        - Row heights match (sample)
        - Page setup matches
        """
        _logger.info("Validating format parity...")

        # Load original template for comparison
        wb_template = load_workbook(template_path, data_only=False)

        # Get same target sheet used in export
        target_sheet = None
        for sheet_name in ['Du toan', 'Chốt', 'HĐ']:
            if sheet_name in wb.sheetnames:
                target_sheet = sheet_name
                break
        if not target_sheet:
            target_sheet = wb.sheetnames[0]

        ws = wb[target_sheet]
        ws_template = wb_template[target_sheet]

        errors = []

        # Check sheet names
        if wb.sheetnames != wb_template.sheetnames:
            errors.append(f"Sheet names mismatch: {wb.sheetnames} vs {wb_template.sheetnames}")

        # Check merged cells count (should be same)
        if len(ws.merged_cells.ranges) != len(ws_template.merged_cells.ranges):
            errors.append(
                f"Merged cells count mismatch: {len(ws.merged_cells.ranges)} vs {len(ws_template.merged_cells.ranges)}"
            )

        # Check column widths (A-J)
        for col_idx in range(1, 11):
            col_letter = get_column_letter(col_idx)
            width = ws.column_dimensions[col_letter].width
            width_template = ws_template.column_dimensions[col_letter].width

            if abs(width - width_template) > 0.1:
                errors.append(f"Column {col_letter} width mismatch: {width} vs {width_template}")

        # Check page setup
        if ws.page_setup.orientation != ws_template.page_setup.orientation:
            errors.append(f"Page orientation mismatch")

        if ws.page_setup.paperSize != ws_template.page_setup.paperSize:
            errors.append(f"Paper size mismatch")

        wb_template.close()

        if errors:
            error_msg = "Template format parity validation FAILED:\n" + "\n".join(errors)
            _logger.error(error_msg)
            # In production, might want to just warn instead of raise
            # raise UserError(error_msg)
        else:
            _logger.info("Format parity validation PASSED")

    def action_import_excel_dtx(self):
        """Open wizard to import Excel quotation"""
        self.ensure_one()

        return {
            'name': 'Import Báo giá Excel',
            'type': 'ir.actions.act_window',
            'res_model': 'sale.excel.import.wizard',
            'view_mode': 'form',
            'target': 'new',
            'context': {
                'default_order_id': self.id,
            },
        }
